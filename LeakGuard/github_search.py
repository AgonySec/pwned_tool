from time import sleep

from openpyxl import Workbook

from .utils import filter_blacklisted_user, filter_sensitive_words, make_github_request, \
    create_sheets, config, default_search_list, save_excel_workbook,read_file

BASE_URLS = {
    'issues': "https://api.github.com/search/issues",
    'code': "https://api.github.com/search/code",
    'repositories': "https://api.github.com/search/repositories"
}

def get_one_page_data(data,sheet,search_type):
    items = data.get('items', [])
    for item in items:
        html_url = item['html_url']
        if search_type == "issues":
            description = item.get('title', '无')
            parts = html_url.split('/')
            full_name = f"{parts[3]}/{parts[4]}/"
        elif search_type == "code":
            description = item['repository'].get('description', '无')
            full_name = item['repository']['full_name']
        else:
            description = item.get('description', '无')
            full_name = item['full_name']

        if filter_sensitive_words(description):
            continue
        if filter_blacklisted_user(html_url):
            continue

        print(f"仓库名: {full_name}, 仓库描述: {description}, 泄露地址: {html_url}")
        sheet.append([full_name, description, html_url])

def get_total_pages(response):
    headers_text = response.headers
    total_pages = 1
    if "Link" in headers_text:
        link = response.headers['Link']
        total_pages = int(link.split('page=')[-1].split('>')[0])
    return total_pages


def fetch_and_process_data(search_type, query, sheet):
    base_url = BASE_URLS[search_type]
    url = f"{base_url}?q={query}&per_page=100&sort=updated"
    response = make_github_request(url)
    if not response:
        return

    total_pages = get_total_pages(response)

    for page in range(1, total_pages + 1):
        url = f"{base_url}?q={query}&per_page=100&page={page}&sort=updated"
        response = make_github_request(url)
        if not response:
            continue
        data = response.json()
        get_one_page_data(data, sheet, search_type)
        sleep(1)



def search_github(query=None,file=None,output_file=None,mode='xlsx'):
    print("开始进行GitHub关键字检测")
    # print_sensitive_words_blacklist_words()
    wb = Workbook()
    if file:
        search_list = read_file(file)
    else:
        search_list = [query] if query else config.get('SearchList', default_search_list)
    sheets = create_sheets(wb, search_list)

    for search_type in BASE_URLS.keys():
        for i, query in enumerate(search_list):
            fetch_and_process_data(search_type, query, sheets[i])
            sheets[i].sheet_properties.tabColor = "9AFF9A"
        sleep(1)

    wb.remove(wb['Sheet'])
    save_excel_workbook(wb, output_file,mode)
