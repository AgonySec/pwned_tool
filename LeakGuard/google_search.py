import re
import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl import Workbook
from .utils import set_socks5_proxy, get_random_user_agent, read_file, save_excel_workbook,save_excel_or_json,get_email_suffixes,create_sheets

# 禁用不安全请求的警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
# google_email_columns=['搜索关键字', '邮箱信息', '站点标题', '发布者', '链接']

# 搜索后缀,返回结果
def fetch_emails(email_suffix,sheet):
    print(f"[*] 正在谷歌搜索邮箱后缀为：{email_suffix} 的邮箱...")
    query = f"{email_suffix}"

    # 设置 SOCKS5 代理
    set_socks5_proxy('localhost', 10808)  # 修改为你实际的 SOCKS5 代理地址和端口
    # 设置 HTTP 代理
    # proxies = {
    #     "http": "http://127.0.0.1:10809"
    #     # "https": "http://127.0.0.1:7890",
    # }

    # 目前测试最大100，后续可改用时间区间来进行最大值搜索
    num = 100
    query = query.replace(' ', '+')
    URL = f"https://google.com/search?q=intext:{query}&sca_esv=d22370baf961194e&sxsrf=ADLYWIJRjgSKAET1MsMdBX705lRoOZwfSA:1718955184914&source=lnt&tbs=li:1&sa=X&ved=2ahUKEwjPxLmJl-yGAxXIrlYBHdbwDWgQpwV6BAgBEBk&biw=1707&bih=898&dpr=1.5&num={num}"

    headers = {"user-agent":get_random_user_agent()}

    # 发送请求
    # resp = requests.get(URL, headers=headers, proxies=proxies, verify=False)
    resp = requests.get(URL, headers=headers, verify=False)

    # print (resp.content)
    # 定义了一个正则表达式, 用于匹配特定后缀的电子邮件地址
    email_pattern = rf'[A-Za-z0-9._%+-]+{email_suffix}'

    # results = []
    seen_results = set() # 用于存储已见过的结果
    if resp.status_code == 200:
        # print("[*] 响应200，内容：:", resp.text)
        # 使用BeautifulSoup解析HTML内容
        soup = BeautifulSoup(resp.content, "html.parser")
        # 查找所有类名为 MjjYud 的div元素
        divs = soup.find_all('div', class_='MjjYud')
        for div in divs:
            # 标题
            title_tag = div.find('h3', class_='LC20lb MBeuO DKV0Md')
            title = title_tag.text if title_tag else 'N/A'
            # 内容简介
            content_tag = div.find('div', class_=lambda value: value and value.startswith("VwiC3b yXK7lf"))
            content = content_tag.text if content_tag else 'N/A'
            # 邮箱地址
            emails = re.findall(email_pattern, content)
            # 链接
            link_tag = div.find('a', href=True)
            link = link_tag['href'] if link_tag else 'N/A'
            # 发布者
            publisher_tag = div.find('span', class_='VuuXrf')
            publisher = publisher_tag.text if publisher_tag else 'N/A'
            keyword = f"{query}"

            # 去重处理！！！！
            for email in emails:
                if email.endswith(email_suffix):
                    result = [keyword, email, title, publisher, link]
                    result_str = str(result)  # 将结果转换为字符串以便去重
                    if result_str not in seen_results:
                        seen_results.add(result_str)
                        sheet.append(result)
                        # print(result)
                        # results.append(result)
        
    
    print("[*] 该邮箱后缀搜索完毕")
    # return results


# 谷歌查找邮箱后缀搜索
def search_google(email_suffix=None, file=None, output_file=None,mode=None):
    wb = Workbook()
    # results=[]
    if email_suffix:
        # 处理单个邮箱后缀
        search_list = [email_suffix]
        # search_list.extend(fetch_emails(email_suffix))
    else:
        # 读取邮箱文件，获取邮箱后缀列表
        emails = read_file(file)
        search_list = get_email_suffixes(emails)
    sheets = create_sheets(wb, search_list, mode='google')

    for i, email_suffix in enumerate(search_list):
        fetch_emails(email_suffix,sheets[i])
        sheets[i].sheet_properties.tabColor = "9AFF9A"

    wb.remove(wb['Sheet'])
    save_excel_workbook(wb, output_file=output_file, key="google", mode=mode)